import os
from datetime import date, datetime
from typing import Any, Dict, List
from openpyxl import load_workbook
import app.constants.result_fields as R

SHEET_NAME = "Eigenbemühungen"


def write_results_to_excels(template_path: str, results: List[Dict[str, Any]], since_date: date, agreed_count: int | None, customer_number: str | None, first_name: str | None, last_name: str | None) -> List[str]:
    """
    Write merged application results into one or more Excel files based on a template.

    Behavior:
    -Loads the given template workbook "Eigenbemühungen"
    -Fills header fields (period, customer number, name, agreed_count)
    -Writes up to 6 entries per file (chunk_size = 6)
    -Each entry occupies 3 rows in the sheet
    -Creates one output file per chunk with a timestamped filename

    Inputs:
    -template_path: path to the Excel template
    -results: list of merged result dicts
    -since_date: start date for the reporting period
    -agreed_count: optional target number written into the sheet
    -customer_number/first_name/last_name: optional personal fields

    Output:
    -List of generated Excel file paths.
    """
    out_files: List[str] = []

    if not results:
        return out_files

    ts = datetime.now().strftime("%d%m%Y_%H%M%S")

    chunk_size = 6
    index = 0
    part = 1

    today = date.today()

    while index < len(results):
        chunk = results[index : index + chunk_size]

        wb = load_workbook(template_path)
        ws = wb[SHEET_NAME]

        period = since_date.strftime(R.DATE_FORMAT) + " - " + today.strftime(R.DATE_FORMAT)
        ws["E4"].value = period

        if customer_number:
            ws["F4"].value = customer_number

        cell = ws["E1"]
        text = str(cell.value or "").strip()
        if text == "Name:" and last_name:
            cell.value = "Name: " + last_name

        cell = ws["F1"]
        text = str(cell.value or "").strip()
        if text == "Vorname:" and first_name:
            cell.value = "Vorname: " + first_name

        if agreed_count is not None:
            ws["B2"].value = agreed_count

        base_row = 7

        for j in range(chunk_size):
            r0 = base_row + j * 3
            ws["A" + str(r0)].value = index + j + 1

        for i, entry in enumerate(chunk):
            r = base_row + i * 3

            employer_name = entry.get(R.EMPLOYER_NAME)
            postal_address = entry.get(R.POSTAL_ADDRESS)

            if employer_name is not None:
                employer_text = employer_name
            else:
                employer_text = ""

            if postal_address is not None and str(postal_address).strip():
                if str(employer_text).strip():
                    employer_text = str(employer_text).strip() + ", " + str(postal_address).strip()
                else:
                    employer_text = str(postal_address).strip()

            ws["B" + str(r)].value = employer_text
            ws["C" + str(r)].value = entry.get(R.CONTACT_PERSON)

            contact_date = entry.get(R.FIRST_CONTACT_DATE)
            role = entry.get(R.APPLIED_POSITION)

            ws["D" + str(r)].value = "am: " + (contact_date if contact_date else "")
            ws["D" + str(r + 1)].value = "wie: Online-Bewerbung"
            ws["D" + str(r + 2)].value = "als: " + (role if role else "")

            ws["F" + str(r)].value = entry.get(R.RESULT)

        out_path = os.path.join(
            os.path.dirname(template_path),
            "table_filled_" + ts + "_" + str(part).zfill(2) + ".xlsx",
        )

        wb.save(out_path)
        out_files.append(out_path)

        index += len(chunk)
        part += 1

    return out_files